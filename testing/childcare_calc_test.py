from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import xlrd
import openpyxl

PATH = "C:\Program Files (x86)\chromedriver.exe"
input_file = "calc_test.xlsx"
output_file = "calc_test.xlsx"
webpage = "http://elenalash.pythonanywhere.com/"


def test_case_one(this_case):
    global error, out_info
    error = ""
    out_info = []
    try:
        driver = webdriver.Chrome(PATH)
        driver.get(webpage)
    except:
        error = "Page could not be accessed"
        return error, out_info
    input_names = ["activ_you", "partner_val", "activ_part", "income", "provider", "daily_fee", "session", "days_per_w"]
    i = 0
    while i < len(input_names):
        if i == 1:
            if this_case[1] == "1":
                driver.find_element_by_name(input_names[i]).click()
                driver.find_element_by_name(input_names[i + 1]).send_keys(this_case[i + 1])
        elif i == 2:
            pass
        else:
            driver.find_element_by_name(input_names[i]).send_keys(this_case[i])
        i += 1
    try:
        calc_button = driver.find_element_by_id("calculate_button")
        calc_button.send_keys(Keys.RETURN)
    except Exception as sub_ex:
        error = sub_ex
        return error, out_info
    out_vars = ["income-out", "hours-out", "rate-out", "fpd-out", "spd-out", "opd-out", "fpw-out", "spw-out", "opw-out",
                "fpa-out", "spa-out", "opa-out"]
    try:
        wait = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, out_vars[0]))
        )
        for var in out_vars:
            out_value = driver.find_element_by_id(var).text
            if out_value == "":
                error = driver.find_element_by_id("input_error").text
                break
            else:
                out_value = char_remover(out_value)
                out_value = str_to_num(out_value)
                out_info.append(out_value)
    except Exception as exs:
        error = driver.find_element_by_id("input_error").text
        return error, out_info
    driver.quit()
    return error, out_info


def str_to_num(item):
    try:
        item = float(item)
        if item % 1 == 0:
            item = int(item)
    except:
        item = item
    return item


def char_remover(string):
    char = ["$", "[ ", " ]", "%", " hours", ","]
    for item in char:
        string = string.replace(item, "")
    return string


wb = xlrd.open_workbook(input_file)
datasheet = wb.sheet_by_name("data")
testCases = []
for i in range(1, datasheet.nrows):
    testCase = []
    for j in range(1, datasheet.ncols):
        value = datasheet.cell_value(i, j)
        value = str_to_num(value)
        testCase.append(str(value))
        j += 1
    testCases.append(testCase)
    i += 1
wb_out = openpyxl.load_workbook(output_file)
ws = wb_out["output"]
start_cell = [2, 3]
for case in testCases:
    error, out_info = test_case_one(case)
    if not error:
        for value in out_info:
            ws.cell(row=testCases.index(case)+start_cell[0], column=out_info.index(value)+start_cell[1], value=value)
    else:
        ws.cell(row=testCases.index(case)+start_cell[0],column=start_cell[1]-1, value=error)
wb_out.save(output_file)
print(f"All done! {len(testCases)} test cases completed.")
